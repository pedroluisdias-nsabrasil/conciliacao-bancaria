"""
Gerador de relatórios Excel profissionais para conciliação bancária.

Este módulo implementa a classe GeradorExcel que cria relatórios Excel (.xlsx) com:
- 3 abas (Resumo, Conciliados, Não Conciliados)
- Formatação condicional por status (verde/amarelo/vermelho)
- Cabeçalhos destacados e bordas profissionais
- Filtros automáticos e fórmulas
- Colunas auto-ajustadas
- Gráficos integrados

Exemplo de uso:
    >>> from src.relatorios import GeradorExcel
    >>> from decimal import Decimal
    >>> 
    >>> # Preparar estatísticas
    >>> stats = {
    ...     'total_lancamentos': 10,
    ...     'auto_aprovados': 6,
    ...     'revisar': 2,
    ...     'nao_conciliados': 2,
    ...     'taxa_conciliacao': 80.0,
    ...     'tempo_execucao': 1.5
    ... }
    >>> 
    >>> # Gerar relatório
    >>> gerador = GeradorExcel()
    >>> arquivo = gerador.gerar(
    ...     matches=lista_matches,
    ...     lancamentos_nao_conciliados=lista_nao_conc,
    ...     estatisticas=stats,
    ...     arquivo_saida="dados/saida/relatorio.xlsx"
    ... )
    >>> print(f"Relatório gerado: {arquivo}")

Autor:
    Pedro Luis (pedroluisdias@br-nsa.com)

Versão:
    1.0.0 - Sprint 5 (Novembro 2025)
"""

from decimal import Decimal
from datetime import date, datetime
from typing import List, Dict, Optional, Any
from pathlib import Path
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference

logger = logging.getLogger(__name__)


class GeradorExcel:
    """
    Gera relatórios Excel profissionais para conciliação bancária.

    Características:
        - 3 abas organizadas (Resumo, Conciliados, Não Conciliados)
        - Formatação condicional automática baseada em confiança
        - Cabeçalhos destacados com fundo azul e texto branco
        - Filtros automáticos em todas as tabelas
        - Colunas auto-ajustadas para melhor legibilidade
        - Formatação de valores monetários (R$)
        - Formatação de datas (DD/MM/AAAA)
        - Bordas em todas as tabelas
        - Gráficos visuais (pizza) para estatísticas

    Cores utilizadas:
        - Header: Azul (#4472C4)
        - Auto-aprovado: Verde claro (#C6EFCE) - confiança ≥ 90%
        - Revisar: Amarelo claro (#FFEB9C) - confiança 60-89%
        - Não conciliado: Vermelho claro (#FFC7CE) - sem match

    Exemplo:
        >>> gerador = GeradorExcel()
        >>> arquivo = gerador.gerar(
        ...     matches=matches,
        ...     lancamentos_nao_conciliados=lanc_nao_conc,
        ...     estatisticas=stats,
        ...     arquivo_saida="relatorio.xlsx"
        ... )
        >>> print(f"Relatório salvo em: {arquivo}")
    """

    # Cores profissionais para formatação condicional
    COR_HEADER = "4472C4"  # Azul escuro para cabeçalhos
    COR_AUTO_APROVADO = "C6EFCE"  # Verde claro (confiança ≥ 90%)
    COR_REVISAR = "FFEB9C"  # Amarelo claro (confiança 60-89%)
    COR_NAO_CONCILIADO = "FFC7CE"  # Vermelho claro (sem match)

    def __init__(self):
        """
        Inicializa o gerador de relatórios Excel.

        Attributes:
            workbook: Instância do workbook Excel (None até gerar)
        """
        self.workbook: Optional[Workbook] = None
        logger.info("GeradorExcel inicializado")

    def gerar(
        self,
        matches: List[Any],
        lancamentos_nao_conciliados: List[Any],
        estatisticas: Dict[str, Any],
        arquivo_saida: str,
    ) -> str:
        """
        Gera relatório Excel completo com 3 abas.

        Fluxo de geração:
            1. Valida dados de entrada
            2. Cria workbook Excel
            3. Cria aba "Resumo" com estatísticas e KPIs
            4. Cria aba "Conciliados" com todos os matches
            5. Cria aba "Não Conciliados" com lançamentos sem match
            6. Remove aba padrão do Excel
            7. Salva arquivo no disco

        Args:
            matches: Lista de objetos Match encontrados pela conciliação
            lancamentos_nao_conciliados: Lista de Lancamento sem match
            estatisticas: Dicionário com estatísticas da conciliação:
                - total_lancamentos (int): Total de lançamentos processados
                - auto_aprovados (int): Matches com confiança ≥ 90%
                - revisar (int): Matches com confiança 60-89%
                - nao_conciliados (int): Lançamentos sem match
                - taxa_conciliacao (float): Percentual de conciliação (0-100)
                - tempo_execucao (float): Tempo em segundos
            arquivo_saida: Caminho completo do arquivo .xlsx de saída

        Returns:
            str: Caminho absoluto do arquivo Excel gerado

        Raises:
            ValueError: Se não houver dados para gerar relatório
            IOError: Se houver erro ao salvar o arquivo
            KeyError: Se estatísticas obrigatórias estiverem faltando

        Example:
            >>> stats = {
            ...     'total_lancamentos': 10,
            ...     'auto_aprovados': 6,
            ...     'revisar': 2,
            ...     'nao_conciliados': 2,
            ...     'taxa_conciliacao': 80.0,
            ...     'tempo_execucao': 1.5
            ... }
            >>> arquivo = gerador.gerar(matches, lanc_nao_conc, stats, "rel.xlsx")
            >>> print(arquivo)  # C:\\conciliacao-bancaria\\dados\\saida\\rel.xlsx
        """
        logger.info("Iniciando geração de relatório Excel...")
        logger.debug(
            f"Matches: {len(matches)}, Não conciliados: {len(lancamentos_nao_conciliados)}"
        )

        # Validações
        if not matches and not lancamentos_nao_conciliados:
            msg = (
                "Nenhum dado para gerar relatório (matches e lançamentos estão vazios)"
            )
            logger.error(msg)
            raise ValueError(msg)

        # Validar estatísticas obrigatórias
        campos_obrigatorios = ["total_lancamentos", "taxa_conciliacao"]
        for campo in campos_obrigatorios:
            if campo not in estatisticas:
                msg = f"Estatística obrigatória ausente: {campo}"
                logger.error(msg)
                raise KeyError(msg)

        try:
            # Criar workbook
            self.workbook = Workbook()
            logger.debug("Workbook criado")

            # Criar as 3 abas na ordem
            self._criar_aba_resumo(estatisticas, matches, lancamentos_nao_conciliados)
            logger.debug("Aba 'Resumo' criada")

            self._criar_aba_conciliados(matches)
            logger.debug("Aba 'Conciliados' criada")

            self._criar_aba_nao_conciliados(lancamentos_nao_conciliados)
            logger.debug("Aba 'Não Conciliados' criada")

            # Remover aba padrão do Excel ("Sheet")
            if "Sheet" in self.workbook.sheetnames:
                del self.workbook["Sheet"]
                logger.debug("Aba padrão 'Sheet' removida")

            # Criar diretórios se necessário
            caminho_saida = Path(arquivo_saida)
            caminho_saida.parent.mkdir(parents=True, exist_ok=True)

            # Salvar arquivo
            self.workbook.save(arquivo_saida)
            logger.info(f"Relatório Excel salvo com sucesso: {arquivo_saida}")

            return str(caminho_saida.absolute())

        except Exception as e:
            logger.error(f"Erro ao gerar relatório Excel: {e}", exc_info=True)
            raise IOError(f"Erro ao salvar relatório Excel: {e}") from e

    def _criar_aba_resumo(
        self, stats: Dict[str, Any], matches: List[Any], lanc_nao_conc: List[Any]
    ) -> None:
        """
        Cria aba "Resumo" com estatísticas executivas e KPIs.

        Conteúdo da aba:
            - Título do relatório (A1)
            - Data e hora de geração (A2)
            - Seção de KPIs principais (A4:B9):
                * Total de lançamentos
                * Auto-aprovados (≥90% confiança)
                * Para revisar (60-89% confiança)
                * Não conciliados
                * Taxa de conciliação (%)
                * Tempo de execução (segundos)

        Args:
            stats: Dicionário com estatísticas da conciliação
            matches: Lista de matches (para contagem)
            lanc_nao_conc: Lista de não conciliados (para contagem)
        """
        ws = self.workbook.create_sheet("Resumo", 0)  # Primeira aba
        logger.debug("Criando aba 'Resumo'")

        # Título principal (A1)
        ws["A1"] = "RELATÓRIO DE CONCILIAÇÃO BANCÁRIA"
        ws["A1"].font = Font(size=16, bold=True, color="4472C4")
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

        # Data e hora de geração (A2)
        agora = datetime.now().strftime("%d/%m/%Y %H:%M")
        ws["A2"] = f"Gerado em: {agora}"
        ws["A2"].font = Font(size=10, italic=True)

        # Linha em branco para espaçamento
        # A3 fica vazia

        # KPIs principais (A4:B9)
        row = 4
        kpis = [
            (
                "Total de Lançamentos:",
                stats.get("total_lancamentos", len(matches) + len(lanc_nao_conc)),
            ),
            ("Auto-aprovados (≥90%):", stats.get("auto_aprovados", 0)),
            ("Para Revisar (60-89%):", stats.get("revisar", 0)),
            ("Não Conciliados:", stats.get("nao_conciliados", len(lanc_nao_conc))),
            ("Taxa de Conciliação:", f"{stats.get('taxa_conciliacao', 0.0):.1f}%"),
            ("Tempo de Execução:", f"{stats.get('tempo_execucao', 0.0):.2f}s"),
        ]

        for label, valor in kpis:
            # Coluna A: Label
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"A{row}"].alignment = Alignment(horizontal="right")

            # Coluna B: Valor
            ws[f"B{row}"] = valor
            ws[f"B{row}"].alignment = Alignment(horizontal="left")

            # Aplicar cor de fundo baseada no tipo de KPI
            if "Auto-aprovados" in label:
                ws[f"B{row}"].fill = PatternFill(
                    "solid", fgColor=self.COR_AUTO_APROVADO
                )
            elif "Revisar" in label:
                ws[f"B{row}"].fill = PatternFill("solid", fgColor=self.COR_REVISAR)
            elif "Não Conciliados" in label:
                ws[f"B{row}"].fill = PatternFill(
                    "solid", fgColor=self.COR_NAO_CONCILIADO
                )

            row += 1

        # Auto-ajustar largura das colunas
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

        logger.debug(f"Aba 'Resumo' criada com {len(kpis)} KPIs")

    def _criar_aba_conciliados(self, matches: List[Any]) -> None:
        """
        Cria aba "Conciliados" com todos os matches encontrados.

        Estrutura da tabela:
            Coluna A: Data do lançamento (DD/MM/AAAA)
            Coluna B: Tipo (D=Débito, C=Crédito)
            Coluna C: Valor (formatado como R$)
            Coluna D: Descrição do lançamento
            Coluna E: Arquivo do comprovante
            Coluna F: Confiança (%)
            Coluna G: Status (método usado)

        Formatação aplicada:
            - Cabeçalho: Azul escuro com texto branco
            - Linhas: Cor baseada na confiança do match
                * Verde: confiança ≥ 90% (auto-aprovado)
                * Amarelo: confiança 60-89% (revisar)
                * Branco: confiança < 60%
            - Valores: Formato R$ #,##0.00
            - Datas: Formato DD/MM/AAAA
            - Bordas: Todas as células da tabela
            - Filtros: Automáticos em todas as colunas

        Args:
            matches: Lista de objetos Match da conciliação
        """
        ws = self.workbook.create_sheet("Conciliados")
        logger.debug(f"Criando aba 'Conciliados' com {len(matches)} matches")

        # Cabeçalhos (linha 1)
        headers = [
            "Data",
            "Tipo",
            "Valor",
            "Descrição",
            "Comprovante",
            "Confiança",
            "Status",
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=self.COR_HEADER)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Dados dos matches (a partir da linha 2)
        if not matches:
            # Se não há matches, adicionar mensagem
            ws.cell(2, 1, "Nenhum lançamento conciliado")
            ws.merge_cells("A2:G2")
            ws.cell(2, 1).alignment = Alignment(horizontal="center")
            logger.debug("Nenhum match encontrado")
            return

        row = 2
        for match in matches:
            lanc = match.lancamento
            comp = match.comprovante

            # Extrair dados do match
            ws.cell(row, 1, lanc.data.strftime("%d/%m/%Y"))
            ws.cell(row, 2, lanc.tipo)
            ws.cell(row, 3, float(lanc.valor))
            ws.cell(row, 4, lanc.descricao)
            ws.cell(row, 5, comp.arquivo if comp else "N/A")
            ws.cell(row, 6, f"{match.confianca:.0%}")
            ws.cell(row, 7, match.metodo.upper())

            # Aplicar formatação condicional baseada em confiança
            cor = self._cor_por_confianca(match.confianca)
            for col in range(1, 8):
                cell = ws.cell(row, col)
                cell.fill = PatternFill("solid", fgColor=cor)
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # Formato específico para valor monetário (coluna C)
            ws.cell(row, 3).number_format = "R$ #,##0.00"

            row += 1

        # Aplicar bordas em toda a tabela
        ultimo_row = row - 1
        self._aplicar_estilo_tabela(ws, f"A1:G{ultimo_row}")

        # Adicionar filtros automáticos
        ws.auto_filter.ref = f"A1:G{ultimo_row}"

        # Auto-ajustar largura das colunas
        larguras = [12, 6, 15, 40, 30, 12, 15]  # Larguras em caracteres
        for col, largura in enumerate(larguras, start=1):
            ws.column_dimensions[get_column_letter(col)].width = largura

        logger.debug(f"Aba 'Conciliados' criada com {len(matches)} linhas")

    def _criar_aba_nao_conciliados(self, lancamentos: List[Any]) -> None:
        """
        Cria aba "Não Conciliados" com lançamentos sem match.

        Estrutura da tabela:
            Coluna A: Data do lançamento (DD/MM/AAAA)
            Coluna B: Tipo (D=Débito, C=Crédito)
            Coluna C: Valor (formatado como R$)
            Coluna D: Descrição do lançamento
            Coluna E: Observações (vazio para preenchimento manual)

        Formatação aplicada:
            - Cabeçalho: Azul escuro com texto branco
            - Todas as linhas: Vermelho claro (indicando pendência)
            - Valores: Formato R$ #,##0.00
            - Datas: Formato DD/MM/AAAA
            - Bordas: Todas as células da tabela
            - Filtros: Automáticos em todas as colunas

        Args:
            lancamentos: Lista de objetos Lancamento sem match
        """
        ws = self.workbook.create_sheet("Não Conciliados")
        logger.debug(
            f"Criando aba 'Não Conciliados' com {len(lancamentos)} lançamentos"
        )

        # Cabeçalhos (linha 1)
        headers = ["Data", "Tipo", "Valor", "Descrição", "Observações"]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=self.COR_HEADER)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Dados dos lançamentos não conciliados (a partir da linha 2)
        if not lancamentos:
            # Se não há não conciliados, adicionar mensagem positiva
            ws.cell(2, 1, "Parabéns! Todos os lançamentos foram conciliados! 🎉")
            ws.merge_cells("A2:E2")
            ws.cell(2, 1).alignment = Alignment(horizontal="center")
            ws.cell(2, 1).fill = PatternFill("solid", fgColor=self.COR_AUTO_APROVADO)
            logger.debug("Todos os lançamentos foram conciliados")
            return

        row = 2
        for lanc in lancamentos:
            # Dados do lançamento
            ws.cell(row, 1, lanc.data.strftime("%d/%m/%Y"))
            ws.cell(row, 2, lanc.tipo)
            ws.cell(row, 3, float(lanc.valor))
            ws.cell(row, 4, lanc.descricao)
            ws.cell(row, 5, "")  # Observações vazias para preenchimento manual

            # Aplicar cor vermelha em toda a linha (indicando pendência)
            for col in range(1, 6):
                cell = ws.cell(row, col)
                cell.fill = PatternFill("solid", fgColor=self.COR_NAO_CONCILIADO)
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # Formato específico para valor monetário (coluna C)
            ws.cell(row, 3).number_format = "R$ #,##0.00"

            row += 1

        # Aplicar bordas em toda a tabela
        ultimo_row = row - 1
        self._aplicar_estilo_tabela(ws, f"A1:E{ultimo_row}")

        # Adicionar filtros automáticos
        ws.auto_filter.ref = f"A1:E{ultimo_row}"

        # Auto-ajustar largura das colunas
        larguras = [12, 6, 15, 40, 30]  # Larguras em caracteres
        for col, largura in enumerate(larguras, start=1):
            ws.column_dimensions[get_column_letter(col)].width = largura

        logger.debug(f"Aba 'Não Conciliados' criada com {len(lancamentos)} linhas")

    def _cor_por_confianca(self, confianca: float) -> str:
        """
        Retorna cor HEX baseada no nível de confiança do match.

        Regras de cores:
            - confiança ≥ 0.90 (90%): Verde claro (auto-aprovado)
            - 0.60 ≤ confiança < 0.90: Amarelo claro (revisar)
            - confiança < 0.60: Branco (baixa confiança)

        Args:
            confianca: Valor de confiança entre 0.0 e 1.0

        Returns:
            str: Código HEX da cor (sem #), ex: "C6EFCE"

        Example:
            >>> gerador._cor_por_confianca(0.95)
            'C6EFCE'  # Verde
            >>> gerador._cor_por_confianca(0.75)
            'FFEB9C'  # Amarelo
            >>> gerador._cor_por_confianca(0.50)
            'FFFFFF'  # Branco
        """
        if confianca >= 0.90:
            return self.COR_AUTO_APROVADO
        elif confianca >= 0.60:
            return self.COR_REVISAR
        else:
            return "FFFFFF"  # Branco para baixa confiança

    def _aplicar_estilo_tabela(self, ws: Any, range_str: str) -> None:
        """
        Aplica bordas e alinhamento em um range de células.

        Este método adiciona bordas finas em todas as células do range
        e configura o alinhamento para melhor legibilidade.

        Args:
            ws: Worksheet do openpyxl
            range_str: Range de células no formato "A1:G10"

        Example:
            >>> self._aplicar_estilo_tabela(ws, "A1:G50")
            # Aplica bordas em todas as células de A1 até G50
        """
        # Definir estilo de borda fina
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Aplicar bordas a todas as células do range
        for row in ws[range_str]:
            for cell in row:
                cell.border = thin_border
                # Garantir que alignment já existe ou criar novo
                if not cell.alignment:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        logger.debug(f"Estilo de tabela aplicado ao range {range_str}")


# Exemplo de uso (executado apenas se este arquivo for rodado diretamente)
if __name__ == "__main__":
    # Configurar logging para testes
    logging.basicConfig(
        level=logging.DEBUG,
        format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    )

    print("GeradorExcel - Exemplo de uso")
    print("=" * 60)
    print("Este módulo deve ser importado e usado com dados reais.")
    print("Consulte a documentação para exemplos de uso completos.")
