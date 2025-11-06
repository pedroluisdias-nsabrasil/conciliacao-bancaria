"""
Router de Relatórios
Gerencia geração e download de relatórios

Autor: Pedro Luis
Data: 06/11/2025
"""

from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse
from pathlib import Path
from datetime import datetime
import logging
import sys

# Adicionar src/ ao path
BASE_DIR = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(BASE_DIR))
sys.path.insert(0, str(BASE_DIR / "src"))

# Imports do código core
from src.relatorios.gerador_excel import GeradorExcel
from api.routers.conciliar import ultimos_resultados

logger = logging.getLogger(__name__)

# ==================== CONFIGURAÇÃO DO ROUTER ====================

router = APIRouter()

# ==================== DIRETÓRIO DE SAÍDA ====================

SAIDA_DIR = BASE_DIR / "dados" / "saida"
SAIDA_DIR.mkdir(parents=True, exist_ok=True)

logger.info(f"📂 Diretório de saída: {SAIDA_DIR}")

# ==================== ENDPOINT: DOWNLOAD DE RELATÓRIO ====================

@router.get("/download")
async def download_relatorio():
    """
    Gera e retorna relatório Excel para download
    
    Returns:
        Arquivo Excel
        
    Raises:
        HTTPException: Se não houver resultados ou erro na geração
    """
    try:
        # Verificar se há resultados
        if not ultimos_resultados:
            raise HTTPException(
                status_code=400,
                detail="Execute a conciliação antes de gerar o relatório"
            )
        
        logger.info("📊 Gerando relatório Excel...")
        
        # Gerar nome do arquivo com timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"relatorio_conciliacao_{timestamp}.xlsx"
        caminho_arquivo = SAIDA_DIR / nome_arquivo
        
        # Obter dados
        matches = ultimos_resultados.get('matches', [])
        nao_conciliados = ultimos_resultados.get('nao_conciliados', [])
        estatisticas = ultimos_resultados.get('estatisticas', {})
        
        # Gerar relatório
        gerador = GeradorExcel()
        
        # TODO: Ajustar chamada do gerador conforme implementação
        # Por enquanto, vamos criar um arquivo simples
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        
        # ===== ABA 1: ESTATÍSTICAS =====
        ws_stats = wb.active
        ws_stats.title = "Estatísticas"
        
        # Cabeçalhos
        headers = [
            ["Métrica", "Valor"],
            ["Total de Lançamentos", estatisticas.get('total_lancamentos', 0)],
            ["Total de Comprovantes", estatisticas.get('total_comprovantes', 0)],
            ["Matches Confirmados", estatisticas.get('total_matches', 0)],
            ["Taxa de Conciliação", f"{estatisticas.get('taxa_conciliacao', 0):.1f}%"],
            ["Confiança Média", f"{estatisticas.get('confianca_media', 0):.1f}%"],
            ["Auto-aprovados", estatisticas.get('auto_aprovados', 0)],
            ["Não Conciliados", estatisticas.get('nao_conciliados', 0)]
        ]
        
        for row_idx, row_data in enumerate(headers, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_stats.cell(row=row_idx, column=col_idx, value=value)
                
                # Estilizar cabeçalho
                if row_idx == 1:
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.font = Font(bold=True, size=12, color="FFFFFF")
                
                cell.alignment = Alignment(horizontal="left")
        
        # Ajustar largura das colunas
        ws_stats.column_dimensions['A'].width = 30
        ws_stats.column_dimensions['B'].width = 20
        
        # ===== ABA 2: MATCHES =====
        ws_matches = wb.create_sheet("Conciliados")
        
        # Cabeçalhos
        ws_matches.append([
            "Data Lançamento",
            "Valor Lançamento",
            "Descrição",
            "Data Comprovante",
            "Valor Comprovante",
            "Confiança (%)",
            "Tipo",
            "Auto-aprovado"
        ])
        
        # Estilizar cabeçalho
        for cell in ws_matches[1]:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        
        # Dados
        for match in matches:
            lanc = match['lancamento']
            comp = match.get('comprovante')
            
            ws_matches.append([
                lanc['data'],
                lanc['valor'],
                lanc['descricao'],
                comp['data'] if comp else '',
                comp['valor'] if comp else '',
                match['confianca'],
                match['tipo'],
                'Sim' if match['auto_aprovado'] else 'Não'
            ])
        
        # Ajustar larguras
        ws_matches.column_dimensions['A'].width = 15
        ws_matches.column_dimensions['B'].width = 15
        ws_matches.column_dimensions['C'].width = 40
        ws_matches.column_dimensions['D'].width = 15
        ws_matches.column_dimensions['E'].width = 15
        ws_matches.column_dimensions['F'].width = 12
        ws_matches.column_dimensions['G'].width = 15
        ws_matches.column_dimensions['H'].width = 12
        
        # ===== ABA 3: NÃO CONCILIADOS =====
        ws_nao_conc = wb.create_sheet("Não Conciliados")
        
        # Cabeçalhos
        ws_nao_conc.append([
            "Data",
            "Valor",
            "Descrição"
        ])
        
        # Estilizar cabeçalho
        for cell in ws_nao_conc[1]:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            cell.font = Font(bold=True, size=11, color="000000")
            cell.alignment = Alignment(horizontal="center")
        
        # Dados
        for lanc in nao_conciliados:
            ws_nao_conc.append([
                lanc['data'],
                lanc['valor'],
                lanc['descricao']
            ])
        
        # Ajustar larguras
        ws_nao_conc.column_dimensions['A'].width = 15
        ws_nao_conc.column_dimensions['B'].width = 15
        ws_nao_conc.column_dimensions['C'].width = 50
        
        # Salvar arquivo
        wb.save(str(caminho_arquivo))
        
        logger.info(f"✅ Relatório gerado: {nome_arquivo}")
        
        # Retornar arquivo para download
        return FileResponse(
            path=str(caminho_arquivo),
            filename=nome_arquivo,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erro ao gerar relatório: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

# ==================== ENDPOINT: LISTAR RELATÓRIOS ====================

@router.get("/listar")
async def listar_relatorios():
    """
    Lista todos os relatórios gerados
    
    Returns:
        JSON com lista de relatórios
    """
    try:
        relatorios = []
        
        for arquivo in SAIDA_DIR.glob("relatorio_conciliacao_*.xlsx"):
            relatorios.append({
                "nome": arquivo.name,
                "tamanho": arquivo.stat().st_size,
                "data_criacao": datetime.fromtimestamp(arquivo.stat().st_mtime).isoformat()
            })
        
        # Ordenar por data (mais recente primeiro)
        relatorios.sort(key=lambda x: x['data_criacao'], reverse=True)
        
        return {
            "status": "ok",
            "total": len(relatorios),
            "relatorios": relatorios
        }
        
    except Exception as e:
        logger.error(f"❌ Erro ao listar relatórios: {e}")
        raise HTTPException(status_code=500, detail=str(e))
