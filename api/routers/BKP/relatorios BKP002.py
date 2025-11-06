"""
Router de Relatórios
Gerencia geração de relatórios Excel e PDF

Autor: Pedro Luis
Data: 06/11/2025
"""

from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse
from pathlib import Path
from typing import Dict, Any
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ==================== CONFIGURAÇÃO DO ROUTER ====================

router = APIRouter()

# ==================== DIRETÓRIOS ====================

BASE_DIR = Path(__file__).resolve().parent.parent.parent
SAIDA_DIR = BASE_DIR / "dados" / "saida"
SAIDA_DIR.mkdir(parents=True, exist_ok=True)

logger.info(f"📂 Diretório de saída: {SAIDA_DIR}")

# ==================== ENDPOINT: GERAR RELATÓRIO EXCEL ====================

@router.get("/download")
async def download_relatorio():
    """
    Gera e retorna relatório Excel da última conciliação
    
    Returns:
        FileResponse com arquivo Excel
        
    Raises:
        HTTPException: Se não houver dados ou erro na geração
    """
    try:
        # Importar aqui para evitar circular import
        from api.routers.conciliar import ultimos_resultados
        
        if not ultimos_resultados:
            raise HTTPException(
                status_code=400,
                detail="Execute a conciliação antes de gerar o relatório"
            )
        
        logger.info("📊 Gerando relatório Excel...")
        
        # Criar nome do arquivo com timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"relatorio_conciliacao_{timestamp}.xlsx"
        caminho_arquivo = SAIDA_DIR / nome_arquivo
        
        # Criar workbook
        wb = Workbook()
        
        # ===== ABA 1: RESUMO =====
        ws_resumo = wb.active
        ws_resumo.title = "Resumo"
        
        # Cabeçalho
        ws_resumo['A1'] = "RELATÓRIO DE CONCILIAÇÃO BANCÁRIA"
        ws_resumo['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws_resumo['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws_resumo['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws_resumo.merge_cells('A1:D1')
        ws_resumo.row_dimensions[1].height = 30
        
        # Data/hora
        ws_resumo['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}"
        ws_resumo['A2'].font = Font(italic=True)
        ws_resumo.merge_cells('A2:D2')
        
        # Estatísticas
        stats = ultimos_resultados['estatisticas']
        
        linha = 4
        estatisticas_labels = [
            ("Total de Lançamentos:", stats['total_lancamentos']),
            ("Total de Comprovantes:", stats['total_comprovantes']),
            ("Matches Encontrados:", stats['total_matches']),
            ("Taxa de Conciliação:", f"{stats['taxa_conciliacao']}%"),
            ("Confiança Média:", f"{stats['confianca_media']}%"),
            ("Auto-aprovados:", stats['auto_aprovados']),
            ("Não Conciliados:", stats['nao_conciliados']),
        ]
        
        for label, valor in estatisticas_labels:
            ws_resumo[f'A{linha}'] = label
            ws_resumo[f'A{linha}'].font = Font(bold=True)
            ws_resumo[f'B{linha}'] = valor
            linha += 1
        
        # Ajustar larguras
        ws_resumo.column_dimensions['A'].width = 30
        ws_resumo.column_dimensions['B'].width = 20
        
        # ===== ABA 2: CONCILIADOS =====
        ws_conciliados = wb.create_sheet("Conciliados")
        
        # Cabeçalhos
        headers = ["Data Lanç.", "Valor", "Descrição", "Comprovante", "Confiança", "Tipo"]
        for col, header in enumerate(headers, 1):
            cell = ws_conciliados.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Dados
        matches = ultimos_resultados['matches']
        for idx, match in enumerate(matches, 2):
            ws_conciliados[f'A{idx}'] = match['lancamento']['data']
            ws_conciliados[f'B{idx}'] = match['lancamento']['valor']
            ws_conciliados[f'B{idx}'].number_format = 'R$ #,##0.00'
            ws_conciliados[f'C{idx}'] = match['lancamento']['descricao']
            
            # Comprovante
            if match['comprovante']:
                comp_info = f"Nº {match['comprovante']['numero']}" if match['comprovante']['numero'] else "S/N"
            else:
                comp_info = "Regra Automática"
            ws_conciliados[f'D{idx}'] = comp_info
            
            ws_conciliados[f'E{idx}'] = f"{match['confianca']}%"
            ws_conciliados[f'F{idx}'] = match['tipo']
            
            # Colorir confiança
            confianca = match['confianca']
            if confianca >= 90:
                ws_conciliados[f'E{idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif confianca >= 70:
                ws_conciliados[f'E{idx}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                ws_conciliados[f'E{idx}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Ajustar larguras
        ws_conciliados.column_dimensions['A'].width = 12
        ws_conciliados.column_dimensions['B'].width = 15
        ws_conciliados.column_dimensions['C'].width = 50
        ws_conciliados.column_dimensions['D'].width = 20
        ws_conciliados.column_dimensions['E'].width = 12
        ws_conciliados.column_dimensions['F'].width = 15
        
        # ===== ABA 3: NÃO CONCILIADOS =====
        ws_nao_conciliados = wb.create_sheet("Não Conciliados")
        
        # Cabeçalhos
        headers_nc = ["Data", "Valor", "Descrição"]
        for col, header in enumerate(headers_nc, 1):
            cell = ws_nao_conciliados.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Dados
        nao_conciliados = ultimos_resultados['nao_conciliados']
        for idx, lanc in enumerate(nao_conciliados, 2):
            ws_nao_conciliados[f'A{idx}'] = lanc['data']
            ws_nao_conciliados[f'B{idx}'] = lanc['valor']
            ws_nao_conciliados[f'B{idx}'].number_format = 'R$ #,##0.00'
            ws_nao_conciliados[f'C{idx}'] = lanc['descricao']
        
        # Ajustar larguras
        ws_nao_conciliados.column_dimensions['A'].width = 12
        ws_nao_conciliados.column_dimensions['B'].width = 15
        ws_nao_conciliados.column_dimensions['C'].width = 50
        
        # Salvar arquivo
        wb.save(caminho_arquivo)
        
        logger.info(f"✅ Relatório gerado: {nome_arquivo}")
        
        # Retornar arquivo para download
        return FileResponse(
            path=caminho_arquivo,
            filename=nome_arquivo,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erro ao gerar relatório: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

# ==================== ENDPOINT: LISTAR RELATÓRIOS ====================

@router.get("/listar")
async def listar_relatorios():
    """
    Lista todos os relatórios disponíveis
    
    Returns:
        JSON com lista de relatórios
    """
    try:
        relatorios = []
        
        for arquivo in SAIDA_DIR.glob("relatorio_*.xlsx"):
            stat = arquivo.stat()
            relatorios.append({
                "nome": arquivo.name,
                "tamanho": f"{stat.st_size / 1024:.1f} KB",
                "data_criacao": datetime.fromtimestamp(stat.st_mtime).strftime("%d/%m/%Y %H:%M:%S"),
                "caminho": str(arquivo)
            })
        
        # Ordenar por data (mais recente primeiro)
        relatorios.sort(key=lambda x: x['data_criacao'], reverse=True)
        
        return {
            "total": len(relatorios),
            "relatorios": relatorios
        }
        
    except Exception as e:
        logger.error(f"❌ Erro ao listar relatórios: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
