"""
Testes automatizados para o GeradorExcel.

Este módulo testa todas as funcionalidades do gerador de relatórios Excel:
- Criação básica de arquivos
- Estrutura de 3 abas
- Formatação condicional
- Valores monetários
- Tratamento de erros
- Validações de dados

Para executar:
    pytest tests/test_relatorios/test_gerador_excel.py -v
    pytest tests/test_relatorios/test_gerador_excel.py -v --cov=src.relatorios.gerador_excel

Autor:
    Pedro Luis (pedroluisdias@br-nsa.com)

Versão:
    1.0.0 - Sprint 5 (Novembro 2025)
"""

from decimal import Decimal
from datetime import date
from pathlib import Path
import pytest

from openpyxl import load_workbook

# Importar o gerador
import sys

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from src.relatorios.gerador_excel import GeradorExcel


# ========== FIXTURES ==========


@pytest.fixture
def lancamentos():
    """
    Cria lançamentos de exemplo para testes.

    Returns:
        Lista de 3 lançamentos simulados com dataclass ou objeto similar
    """
    from dataclasses import dataclass

    @dataclass
    class Lancamento:
        data: date
        tipo: str
        valor: Decimal
        descricao: str

    return [
        Lancamento(
            data=date(2025, 11, 1),
            tipo="D",
            valor=Decimal("150.00"),
            descricao="COMPRA MERCADO",
        ),
        Lancamento(
            data=date(2025, 11, 2),
            tipo="C",
            valor=Decimal("500.00"),
            descricao="PIX RECEBIDO",
        ),
        Lancamento(
            data=date(2025, 11, 3),
            tipo="D",
            valor=Decimal("75.50"),
            descricao="TARIFA BANCARIA",
        ),
    ]


@pytest.fixture
def comprovantes():
    """
    Cria comprovantes de exemplo para testes.

    Returns:
        Lista de 2 comprovantes simulados
    """
    from dataclasses import dataclass

    @dataclass
    class Comprovante:
        arquivo: str
        data: date
        valor: Decimal

    return [
        Comprovante(
            arquivo="comprovante1.pdf", data=date(2025, 11, 1), valor=Decimal("150.00")
        ),
        Comprovante(
            arquivo="comprovante2.pdf", data=date(2025, 11, 2), valor=Decimal("500.00")
        ),
    ]


@pytest.fixture
def matches(lancamentos, comprovantes):
    """
    Cria matches de exemplo para testes.

    Returns:
        Lista de 2 matches simulados com diferentes níveis de confiança
    """
    from dataclasses import dataclass

    @dataclass
    class Match:
        lancamento: object
        comprovante: object
        confianca: float
        metodo: str

    return [
        Match(
            lancamento=lancamentos[0],
            comprovante=comprovantes[0],
            confianca=0.95,  # Auto-aprovado
            metodo="exato",
        ),
        Match(
            lancamento=lancamentos[1],
            comprovante=comprovantes[1],
            confianca=0.75,  # Revisar
            metodo="fuzzy",
        ),
    ]


@pytest.fixture
def estatisticas():
    """
    Cria estatísticas de exemplo para testes.

    Returns:
        Dicionário com estatísticas simuladas
    """
    return {
        "total_lancamentos": 3,
        "auto_aprovados": 1,
        "revisar": 1,
        "nao_conciliados": 1,
        "taxa_conciliacao": 66.7,
        "tempo_execucao": 1.5,
    }


# ========== TESTES BÁSICOS ==========


def test_criar_gerador_excel():
    """Testa criação básica do GeradorExcel."""
    gerador = GeradorExcel()

    assert gerador is not None
    assert gerador.workbook is None  # Workbook só é criado ao gerar
    assert gerador.COR_HEADER == "4472C4"
    assert gerador.COR_AUTO_APROVADO == "C6EFCE"
    assert gerador.COR_REVISAR == "FFEB9C"
    assert gerador.COR_NAO_CONCILIADO == "FFC7CE"


def test_criar_excel_basico(matches, lancamentos, estatisticas, tmp_path):
    """Testa geração básica de arquivo Excel."""
    gerador = GeradorExcel()
    arquivo = tmp_path / "test.xlsx"

    resultado = gerador.gerar(
        matches=matches,
        lancamentos_nao_conciliados=[lancamentos[2]],
        estatisticas=estatisticas,
        arquivo_saida=str(arquivo),
    )

    # Verificar que arquivo foi criado
    assert Path(resultado).exists()
    assert Path(resultado).stat().st_size > 0
    assert str(arquivo) in resultado


def test_excel_tem_3_abas(matches, lancamentos, estatisticas, tmp_path):
    """Testa que Excel tem 3 abas corretas."""
    gerador = GeradorExcel()
    arquivo = tmp_path / "test.xlsx"

    gerador.gerar(
        matches=matches,
        lancamentos_nao_conciliados=[lancamentos[2]],
        estatisticas=estatisticas,
        arquivo_saida=str(arquivo),
    )

    # Carregar arquivo e verificar abas
    wb = load_workbook(arquivo)
    assert len(wb.sheetnames) == 3
    assert "Resumo" in wb.sheetnames
    assert "Conciliados" in wb.sheetnames
    assert "Não Conciliados" in wb.sheetnames

    # Verificar ordem das abas
    assert wb.sheetnames[0] == "Resumo"
    assert wb.sheetnames[1] == "Conciliados"
    assert wb.sheetnames[2] == "Não Conciliados"


# ========== TESTES DA ABA RESUMO ==========


def test_aba_resumo_tem_titulo(matches, lancamentos, estatisticas, tmp_path):
    """Testa que aba Resumo tem título correto."""
    gerador = GeradorExcel()
    arquivo = tmp_path / "test.xlsx"

    gerador.gerar(
        matches=matches,
        lancamentos_nao_conciliados=[lancamentos[2]],
        estatisticas=estatisticas,
        arquivo_saida=str(arquivo),
    )

    wb = load_workbook(arquivo)
    ws = wb["Resumo"]

    # Verificar título
    assert "CONCILIAÇÃO" in ws["A1"].value.upper()
    assert "BANCÁRIA" in ws["A1"].value.upper() or "BANCARIA" in ws["A1"].value.upper()

    # Verificar formatação do título
    assert ws["A1"].font.size == 16
    assert ws["A1"].font.bold is True


def test_aba_resumo_tem_kpis(matches, lancamentos, estatisticas, tmp_path):
    """Testa que aba Resumo tem KPIs corretos."""
    gerador = GeradorExcel()
    arquivo = tmp_path / "test.xlsx"

    gerador.gerar(
        matches=matches,
        lancamentos_nao_conciliados=[lancamentos[2]],
        estatisticas=estatisticas,
        arquivo_saida=str(arquivo),
    )

    wb = load_workbook(arquivo)
    ws = wb["Resumo"]

    # Verificar que KPIs estão presentes
    # Buscar valores nas células B4 a B9 (onde ficam os valores dos KPIs)
    valores_celulas = [ws[f"B{i}"].value for i in range(4, 10)]

    # Verificar que estatísticas estão presentes
    assert 3 in valores_celulas or "3" in str(valores_celulas)  # Total lançamentos
    assert any(
        "66" in str(v) or "67" in str(v) for v in valores_celulas
    )  # Taxa conciliação


def test_aba_resumo_tem_cores_kpis(matches, lancamentos, estatisticas, tmp_path):
    """Testa que KPIs têm cores corretas na aba Resumo."""
    gerador = GeradorExcel()
    arquivo = tmp_path / "test.xlsx"

    gerador.gerar(
        matches=matches,
        lancamentos_nao_conciliados=[lancamentos[2]],
        estatisticas=estatisticas,
        arquivo_saida=str(arquivo),
    )

    wb = load_workbook(arquivo)
    ws = wb["Resumo"]

    # Verificar que pelo menos uma célula tem cor de fundo
    # (auto-aprovados, revisar, ou não conciliados)
    cores_encontradas = []
    for i in range(4, 10):
        cor = ws[f"B{i}"].fill.fgColor
        if cor and hasattr(cor, "rgb"):
            cores_encontradas.append(cor.rgb)

    assert len(cores_encontradas) > 0  # Pelo menos uma célula colorida


# ========== TESTES DA ABA CONCILIADOS ==========


def test_aba_conciliados_tem_headers(matches, lancamentos, estatisticas, tmp_path):
    """Testa que aba Conciliados tem cabeçalhos corretos."""
    gerador = GeradorExcel()
    arquivo = tmp_path / "test.xlsx"

    gerador.gerar(
        matches=matches,
        lancamentos_nao_conciliados=[lancamentos[2]],
        estatisticas=estatisticas,
        arquivo_saida=str(arquivo),
    )

    wb = load_workbook(arquivo)
    ws = wb["Conciliados"]

    # Verificar cabeçalhos
    headers_esperados = [
        "Data",
        "Tipo",
        "Valor",
        "Descrição",
        "Comprovante",
        "Confiança",
        "Status",
    ]

    for col, header_esperado in enumerate(headers_esperados, start=1):
        header_real = ws.cell(1, col).value
        assert (
            header_real == header_esperado
        ), f"Coluna {col}: esperado '{header_esperado}', obtido '{header_real}'"

    # Verificar formatação do cabeçalho
    assert ws.cell(1, 1).font.bold is True
    assert ws.cell(1, 1).font.color.rgb in ["FFFFFF", "00FFFFFF"]  # Texto branco


def test_formatacao_condicional_cores(matches, lancamentos, estatisticas, tmp_path):
    """Testa cores da formatação condicional em Conciliados."""
    gerador = GeradorExcel()
    arquivo = tmp_path / "test.xlsx"

    gerador.gerar(
        matches=matches,
        lancamentos_nao_conciliados=[lancamentos[2]],
        estatisticas=estatisticas,
        arquivo_saida=str(arquivo),
    )

    wb = load_workbook(arquivo)
    ws = wb["Conciliados"]

    # Verificar linha 2 (match com confiança 0.95 - deve ser verde)
    cor_row2 = ws["A2"].fill.fgColor
    if hasattr(cor_row2, "rgb"):
        assert cor_row2.rgb in [gerador.COR_AUTO_APROVADO, "FFC6EFCE", "00C6EFCE"]

    # Verificar linha 3 (match com confiança 0.75 - deve ser amarelo)
    cor_row3 = ws["A3"].fill.fgColor
    if hasattr(cor_row3, "rgb"):
        assert cor_row3.rgb in [gerador.COR_REVISAR, "FFFFEB9C", "00FFEB9C"]


def test_valores_monetarios_formatados(matches, lancamentos, estatisticas, tmp_path):
    """Testa formatação de valores monetários."""
    gerador = GeradorExcel()
    arquivo = tmp_path / "test.xlsx"

    gerador.gerar(
        matches=matches,
        lancamentos_nao_conciliados=[lancamentos[2]],
        estatisticas=estatisticas,
        arquivo_saida=str(arquivo),
    )

    wb = load_workbook(arquivo)
    ws = wb["Conciliados"]

    # Coluna C (valor) deve ter formato moeda
    formato_valor = ws["C2"].number_format
    assert "R$" in formato_valor or "#,##0.00" in formato_valor


def test_filtros_automaticos(matches, lancamentos, estatisticas, tmp_path):
    """Testa que filtros automáticos estão habilitados."""
    gerador = GeradorExcel()
    arquivo = tmp_path / "test.xlsx"

    gerador.gerar(
        matches=matches,
        lancamentos_nao_conciliados=[lancamentos[2]],
        estatisticas=estatisticas,
        arquivo_saida=str(arquivo),
    )

    wb = load_workbook(arquivo)
    ws = wb["Conciliados"]

    # Verificar que auto_filter está definido
    assert ws.auto_filter.ref is not None
    assert "A1" in ws.auto_filter.ref


# ========== TESTES DA ABA NÃO CONCILIADOS ==========


def test_aba_nao_conciliados_vermelha(matches, lancamentos, estatisticas, tmp_path):
    """Testa que não conciliados ficam vermelhos."""
    gerador = GeradorExcel()
    arquivo = tmp_path / "test.xlsx"

    gerador.gerar(
        matches=matches,
        lancamentos_nao_conciliados=[lancamentos[2]],
        estatisticas=estatisticas,
        arquivo_saida=str(arquivo),
    )

    wb = load_workbook(arquivo)
    ws = wb["Não Conciliados"]

    # Verificar cor vermelha na primeira linha de dados
    cor = ws["A2"].fill.fgColor
    if hasattr(cor, "rgb"):
        assert cor.rgb in [gerador.COR_NAO_CONCILIADO, "FFFFC7CE", "00FFC7CE"]


def test_aba_nao_conciliados_headers(matches, lancamentos, estatisticas, tmp_path):
    """Testa cabeçalhos da aba Não Conciliados."""
    gerador = GeradorExcel()
    arquivo = tmp_path / "test.xlsx"

    gerador.gerar(
        matches=matches,
        lancamentos_nao_conciliados=[lancamentos[2]],
        estatisticas=estatisticas,
        arquivo_saida=str(arquivo),
    )

    wb = load_workbook(arquivo)
    ws = wb["Não Conciliados"]

    # Verificar cabeçalhos
    headers_esperados = ["Data", "Tipo", "Valor", "Descrição", "Observações"]

    for col, header_esperado in enumerate(headers_esperados, start=1):
        header_real = ws.cell(1, col).value
        assert header_real == header_esperado


def test_nao_conciliados_mensagem_quando_vazio(matches, estatisticas, tmp_path):
    """Testa mensagem quando não há lançamentos não conciliados."""
    gerador = GeradorExcel()
    arquivo = tmp_path / "test.xlsx"

    gerador.gerar(
        matches=matches,
        lancamentos_nao_conciliados=[],  # Lista vazia
        estatisticas=estatisticas,
        arquivo_saida=str(arquivo),
    )

    wb = load_workbook(arquivo)
    ws = wb["Não Conciliados"]

    # Verificar mensagem positiva
    mensagem = ws["A2"].value
    assert "Parabéns" in mensagem or "conciliados" in mensagem


# ========== TESTES DE VALIDAÇÃO E ERROS ==========


def test_erro_dados_vazios():
    """Testa erro quando não há dados."""
    gerador = GeradorExcel()

    with pytest.raises(ValueError, match="Nenhum dado"):
        gerador.gerar(
            matches=[],
            lancamentos_nao_conciliados=[],
            estatisticas={},
            arquivo_saida="test.xlsx",
        )


def test_erro_estatisticas_faltando(matches, lancamentos):
    """Testa erro quando estatísticas obrigatórias estão faltando."""
    gerador = GeradorExcel()

    with pytest.raises(KeyError, match="obrigatória"):
        gerador.gerar(
            matches=matches,
            lancamentos_nao_conciliados=[lancamentos[2]],
            estatisticas={},  # Estatísticas vazias
            arquivo_saida="test.xlsx",
        )


def test_cria_diretorios_se_necessario(matches, lancamentos, estatisticas, tmp_path):
    """Testa criação de diretórios."""
    gerador = GeradorExcel()
    arquivo = tmp_path / "subdir" / "subdir2" / "test.xlsx"

    gerador.gerar(
        matches=matches,
        lancamentos_nao_conciliados=[lancamentos[2]],
        estatisticas=estatisticas,
        arquivo_saida=str(arquivo),
    )

    assert arquivo.exists()
    assert arquivo.parent.exists()


# ========== TESTES DE MÉTODOS AUXILIARES ==========


def test_cor_por_confianca():
    """Testa método _cor_por_confianca."""
    gerador = GeradorExcel()

    # Auto-aprovado (≥ 0.90)
    assert gerador._cor_por_confianca(0.95) == gerador.COR_AUTO_APROVADO
    assert gerador._cor_por_confianca(0.90) == gerador.COR_AUTO_APROVADO

    # Revisar (0.60 a 0.89)
    assert gerador._cor_por_confianca(0.75) == gerador.COR_REVISAR
    assert gerador._cor_por_confianca(0.60) == gerador.COR_REVISAR

    # Baixa confiança (< 0.60)
    assert gerador._cor_por_confianca(0.50) == "FFFFFF"
    assert gerador._cor_por_confianca(0.10) == "FFFFFF"


def test_larguras_colunas_ajustadas(matches, lancamentos, estatisticas, tmp_path):
    """Testa que colunas têm largura ajustada."""
    gerador = GeradorExcel()
    arquivo = tmp_path / "test.xlsx"

    gerador.gerar(
        matches=matches,
        lancamentos_nao_conciliados=[lancamentos[2]],
        estatisticas=estatisticas,
        arquivo_saida=str(arquivo),
    )

    wb = load_workbook(arquivo)
    ws = wb["Conciliados"]

    # Verificar que colunas têm largura > 0
    assert ws.column_dimensions["A"].width > 0
    assert ws.column_dimensions["D"].width > 0  # Coluna Descrição deve ser mais larga


def test_bordas_aplicadas(matches, lancamentos, estatisticas, tmp_path):
    """Testa que bordas estão aplicadas nas tabelas."""
    gerador = GeradorExcel()
    arquivo = tmp_path / "test.xlsx"

    gerador.gerar(
        matches=matches,
        lancamentos_nao_conciliados=[lancamentos[2]],
        estatisticas=estatisticas,
        arquivo_saida=str(arquivo),
    )

    wb = load_workbook(arquivo)
    ws = wb["Conciliados"]

    # Verificar que célula tem borda
    cell = ws["A2"]
    assert cell.border is not None
    assert cell.border.left.style is not None


# ========== TESTE DE INTEGRAÇÃO ==========


def test_integracao_completa(matches, lancamentos, estatisticas, tmp_path):
    """Teste de integração completo do GeradorExcel."""
    gerador = GeradorExcel()
    arquivo = tmp_path / "relatorio_completo.xlsx"

    # Gerar relatório completo
    resultado = gerador.gerar(
        matches=matches,
        lancamentos_nao_conciliados=[lancamentos[2]],
        estatisticas=estatisticas,
        arquivo_saida=str(arquivo),
    )

    # Verificações gerais
    assert Path(resultado).exists()
    assert Path(resultado).stat().st_size > 5000  # Arquivo deve ter tamanho razoável

    # Carregar e verificar estrutura
    wb = load_workbook(arquivo)

    # 3 abas
    assert len(wb.sheetnames) == 3

    # Aba Resumo tem dados
    ws_resumo = wb["Resumo"]
    assert ws_resumo["A1"].value is not None
    assert ws_resumo["B4"].value is not None

    # Aba Conciliados tem dados
    ws_conc = wb["Conciliados"]
    assert ws_conc["A1"].value == "Data"
    assert ws_conc["A2"].value is not None  # Primeira linha de dados

    # Aba Não Conciliados tem dados
    ws_nao_conc = wb["Não Conciliados"]
    assert ws_nao_conc["A1"].value == "Data"
    assert ws_nao_conc["A2"].value is not None


# ========== EXECUÇÃO DOS TESTES ==========

if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
